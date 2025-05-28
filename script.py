import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from datetime import datetime

st.title("📦 Daily Sales Updater for Chandra Mohan")

uploaded_file = st.file_uploader("Upload the Excel File", type=["xlsx"])

if uploaded_file:
    xls = pd.ExcelFile(uploaded_file)

    if 'SPECILA' not in xls.sheet_names or 'total' not in xls.sheet_names:
        st.error("Required sheets 'SPECILA' or 'total' not found.")
        st.stop()

    df_specila = xls.parse('SPECILA', header=[0,1])
    df_specila.columns = [' '.join([str(i).strip() for i in col if i]).strip() for col in df_specila.columns]

    # Use exact column names based on your data
    company_col = "Name of the Company Unnamed: 1_level_1"
    variety_col = "Name of the Hybrid Unnamed: 2_level_1"
    stock_col = "Chandra Mohan Stock positioned as on date"
    sales_col = "Chandra Mohan Sales as on date"

    # Check columns existence
    for c in [company_col, variety_col, stock_col, sales_col]:
        if c not in df_specila.columns:
            st.error(f"Column '{c}' not found in 'SPECILA' sheet.")
            st.stop()

    companies = df_specila[company_col].dropna().unique().tolist()
    selected_company = st.selectbox("Select Company", companies)

    varieties = df_specila[df_specila[company_col] == selected_company][variety_col].dropna().unique().tolist()
    selected_variety = st.selectbox("Select Variety", varieties)

    update_options = ["Stock", "Sales"]
    selected_updates = st.multiselect("Select what to update", update_options, default=update_options)

    if not selected_updates:
        st.warning("Please select at least one item to update.")
        st.stop()

    input_values = {}
    if "Stock" in selected_updates:
        input_values['Stock'] = st.number_input(f"Enter amount to add to '{stock_col}'", min_value=0, step=1)
    if "Sales" in selected_updates:
        input_values['Sales'] = st.number_input(f"Enter amount to add to '{sales_col}'", min_value=0, step=1)

    if st.button("Update Sales/Stock"):
        wb = load_workbook(uploaded_file)
        ws = wb['SPECILA']

        # Find dataframe row index for selected company & hybrid
        row_idx = df_specila[
            (df_specila[company_col] == selected_company) & 
            (df_specila[variety_col] == selected_variety)
        ].index

        if row_idx.empty:
            st.error("Selected company and variety not found.")
            st.stop()

        excel_row = row_idx[0] + 3  # header 2 rows + 1 for 1-based index

        # Map col names to excel col number
        col_letters = {col: idx + 1 for idx, col in enumerate(df_specila.columns)}

        for key, val in input_values.items():
            if val == 0:
                continue
            col_name = stock_col if key == "Stock" else sales_col
            col_idx = col_letters.get(col_name)
            if not col_idx:
                st.warning(f"Column '{col_name}' not found in sheet.")
                continue

            cell = ws.cell(row=excel_row, column=col_idx)
            old_val = cell.value if cell.value is not None else 0
            try:
                new_val = float(old_val) + float(val)
            except Exception:
                new_val = val
            cell.value = new_val

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        today_str = datetime.today().strftime("%Y-%m-%d")
        file_name = f"updated_sales_{today_str}.xlsx"

        st.success("File updated successfully!")
        st.download_button(
            label="Download Updated Excel (Check Downloads folder on your phone)",
            data=output,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
